import logging
import tkinter 
from tkinter import messagebox
import shutil
import os
import openpyxl
from openpyxl.styles import NamedStyle
from datetime import datetime
import PIL

logger = logging.getLogger('scalewiz')


def export_report(self, project):
        logger.info(f"Beginning export of {project.name.get()}")
        # save the report
        # self.save()
        # logger.debug(f"{project.name.get()} was saved to its json file")

        template = project.template.get()
        if not os.path.isfile(template):
            msg = f"No valid template file found at {template}"
            messagebox.showinfo(self, message=msg)
            logger.error(msg)
            return

        out = f"{project.numbers.get().replace(' ', '')} {project.name.get()} - CaCO3 Scale Block Analysis.xlsx"
        out = os.path.join(os.path.dirname(project.path.get()), out)
        shutil.copyfile(project.template.get(), out)

        def ret_num(str) -> int:
            str = str.replace(",", "")
            if str == "":
                str = 0
            return round(float(str))

        wb = openpyxl.load_workbook(out)
        ws = wb.active

        img = PIL.Image.open(project.plot.get())
        img = img.resize((667, 257))
        img_path = project.plot.get()[:-4]
        img_path += "- temp.png"
        img.save(img_path)
        img = openpyxl.drawing.image.Image(img_path)
        img.anchor = 'A28'
        ws._images[1] = img

        # select the blanks
        blanks = []
        for test in project.tests:
            if test.isBlank.get() and test.includeOnRep.get():
                blanks.append(test)
       
        # report template supports 2 blanks currently
        if len(blanks) > 2:
            blanks = blanks[1:2]
        blank_times = []
        for blank in blanks:
            blank_times.append(len(blank.readings * project.interval.get()))

        # select the trials
        trials = []
        for test in project.tests:
            if test.includeOnRep.get() and not test.isBlank.get():
                trials.append(test)

        # report template supports 10 trials currently
        if len(trials) > 10:
            trials = trials[1:10]

        chemicals = [trial.chemical.get() for trial in trials]
        rates = [trial.rate.get() for trial in trials]
        results = [trial.result.get() for trial in trials]
        durations = [round(len(trial.readings) * project.interval.get() / 60, 2) for trial in trials]
        baseline = project.baseline.get()
        ylim = project.limitPSI.get()
        max_psis = [trial.maxPSI.get() for trial in trials]
        clarities = [trial.clarity.get() for trial in trials]

        brine_comp = f"Synthetic Field Brine, Chlorides = {project.chlorides.get():,} mg/L"
        if project.bicarbsIncreased.get():
            brine_comp += f" (Bicarbs increased to {project.bicarbs.get():,} mg/L)"
        
        
        
        ws['D12'] = brine_comp
        ws['D10'] = project.temperature.get()
        ws['C4'] = project.customer.get()
        ws['C5'] = project.submittedBy.get()
        ws['C6'] = project.productionCo.get()
        ws['C7'] = project.sample.get()
        if not '-' in project.numbers.get() and project.numbers.get != '':
            # if the analysis no is just one #, excel will show an error
            # for having a number formatted as text
            try:
                ws['I4'] = int(project.numbers.get())
            except Exception:
                pass
        else:
            # otherwise "# - #" is a fine text value for the cell
            ws['I4'] = project.numbers.get()

        try:
            ws['I5'] = datetime.strptime(project.sampleDate.get(), '%m/%d/%Y')
            ws['I6'] = datetime.strptime(project.recDate.get(), '%m/%d/%Y')
            ws['I7'] = datetime.strptime(project.compDate.get(), '%m/%d/%Y')
        except Exception:
            pass
        
        ws['D11'] = project.baseline.get()
        ws['G16'] = project.limitPSI.get()
        ws['A29'] = project.analyst.get()

        blank_time_cells = [f"E{i}" for i in range(16, 18)]
        chem_name_cells = [f"A{i}" for i in range(19, 27)]
        chem_conc_cells = [f"D{i}" for i in range(19, 27)]
        duration_cells = [f"E{i}" for i in range(19, 27)]
        max_psi_cells = [f"G{i}" for i in range(19, 27)]
        protection_cells = [f"H{i}" for i in range(19, 27)]
        clarity_cells = [f"J{i}" for i in range(19, 27)]

        for (cell, blank_time) in zip(blank_time_cells, blank_times):
            ws[cell] = round(blank_time / 60, 2)
        for (cell, chemical) in zip(chem_name_cells, chemicals):
            ws[cell] = f"{chemical}"
        for (cell, rate) in zip(chem_conc_cells, rates):
            ws[cell] = rate
        for (cell, duration) in zip(duration_cells, durations):
            ws[cell] = round(duration, 2)
        for (cell, psi) in zip(max_psi_cells, max_psis):
            if psi > project.limitPSI.get():
                psi = project.limitPSI.get()
            ws[cell] = psi
        for (cell, result) in zip(protection_cells, results):
            # limit the results on the report to 100%
            if result > 1:
                result = float(1)
            ws[cell] = result
        for(cell, clarity) in zip(clarity_cells, clarities):
            ws[cell] = clarity

        rows_with_data = [16, 17, *range(19, 27)]  # where the data is
        hide_rows = []  # rows we want to hide
        resize_rows = []  # rows we want to resize

        for i in rows_with_data:
            if ws[f'A{i}'].value is None:  # if the cell is empty
                hide_rows.append(i)  # add it to the list of rows to hide
            else:
                resize_rows.append(i)  # add it to the list of rows to reszie

        row_height = 200 / len(resize_rows)  # we have ~200px to work with total
        if row_height >= 30:  # we don't want any rows bigger than this
            row_height = 30
        for row in resize_rows:  # this does the resizing
            ws.row_dimensions[row].height = row_height
        for row in hide_rows:  # this hides the empty rows
            ws.row_dimensions[row].hidden = True

        wb.save(filename=out)
        os.remove(img_path)
        messagebox.showinfo(
            parent=self,
            message=f"Report exported to\n{out}"
        )
